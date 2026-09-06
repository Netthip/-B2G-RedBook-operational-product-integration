#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_registry.py — DRAFT v0.1 (รอ Gift อนุมัติ schema ก่อนถือเป็นทะเบียนจริง)

สร้าง dataset registry/manifest ของไฟล์ Excel ในโปรเจกต์ โดย
  * ไม่แก้ / ไม่ย้าย / ไม่ rename ไฟล์ต้นฉบับใด ๆ  (เปิดแบบ read_only เท่านั้น ไม่มี save())
  * ไม่ใช้ "ชื่อไฟล์" เป็นตัวตัดสิน identity — ชื่อไฟล์ถูกบันทึกไว้ในฟิลด์ *_from_filename
    ซึ่งถือเป็น "ค่าที่ประกาศ" (declared) เท่านั้น ส่วนค่าที่ใช้จริงมาจาก *_from_content
  * ทุกค่าที่ดึงจากเนื้อไฟล์ต้องมี evidence pointer (sheet + cell + ข้อความดิบ)

ผลลัพธ์ 3 ไฟล์ใน ../registry/
  registry.internal.json  — ครบทุกฟิลด์ (มี path + ชื่อหน่วยงาน) → LOCAL ONLY ห้ามขึ้น repo
  registry.internal.csv   — มุมมองตารางของไฟล์ข้างบน            → LOCAL ONLY
  registry.public.csv     — ฉบับ sanitized (pseudonym A1/A2…)   → เผยแพร่ได้

อัตลักษณ์จริง (ตารางชื่อหน่วยงาน→รหัส · ชื่อไฟล์ mapping donor · รายชื่อหน่วยงานใน scope)
ไม่ได้ฝังในโค้ด แต่อ่านจากไฟล์ local ที่ไม่อยู่ใน repo:
    ตั้งค่าผ่าน env P1REG_LOCAL_CONFIG หรือวางไว้ที่ ../_local/local_identity.json
ถ้าไม่มีไฟล์นั้น สคริปต์ยังรันได้ แต่จะ resolve รหัสหน่วยงานไม่ได้ (unresolved)

การใช้งาน:
    python build_registry.py        (รันจาก project root)
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import sys
from datetime import datetime, timezone

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
OUT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "registry"))
SCHEMA_VERSION = "p1-dataset-registry/0.1-draft"

# ---------------------------------------------------------------------------
# อัตลักษณ์จริงถูกแยกออกไปไว้ในไฟล์ local ที่ไม่อยู่ใน repo (ดู docstring)
# ---------------------------------------------------------------------------
_DEFAULT_LOCAL = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "..", "_local", "local_identity.json")
_LOCAL_PATH = os.environ.get("P1REG_LOCAL_CONFIG", _DEFAULT_LOCAL)
try:
    with open(_LOCAL_PATH, encoding="utf-8") as _f:
        _LOCAL = json.load(_f)
except (OSError, ValueError):
    _LOCAL = {}
    print("[warn] ไม่พบไฟล์อัตลักษณ์ local — จะ resolve รหัสหน่วยงานไม่ได้", file=sys.stderr)

IN_SCOPE_AGENCIES = set(_LOCAL.get("in_scope_agencies", []))
MAPPING_DONOR_BASENAME = _LOCAL.get("mapping_donor_basename")
MAPPING_DONOR_EVIDENCE = _LOCAL.get("mapping_donor_evidence", "")
AGENCY_NAME_TO_CODE = _LOCAL.get("agency_name_to_code", {})

STAGE_ENUM = {
    "2.3": "req_operating_unit_2_3",
    "2.5": "req_ministry_2_5",
    "3.1": "bb_officer_3_1",
}


# ----------------------------------------------------------------------------- hashing
def file_bytes_sha256(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def semantic_sha256(wb) -> tuple[str, int]:
    """
    แฮชของ "ค่าที่อ่านได้" หลัง normalize — ใช้จับกรณี 'ข้อมูลเดียวกันแต่ save ใหม่จนไบต์ต่าง'

    🔴 ไม่ใช่ reproducibility hash และห้ามใช้อ้างว่า 'รันซ้ำได้ค่าเดิม'
       ครอบคลุมเฉพาะ: ชื่อชีต + พิกัดเซลล์ + ค่าที่ data_only อ่านได้
       ไม่ครอบคลุม: สูตร, รูปแบบ, merged cells, แถว/คอลัมน์ที่ซ่อน, metadata
    """
    h = hashlib.sha256()
    n_cells = 0
    for name in wb.sheetnames:
        h.update(("\x1fSHEET\x1f" + name).encode("utf-8"))
        ws = wb[name]
        for row in ws.iter_rows(values_only=False):
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, float):
                    s = repr(round(v, 6))
                else:
                    s = str(v).strip()
                    if not s:
                        continue
                h.update(("\x1f" + c.coordinate + "\x1f" + s).encode("utf-8"))
                n_cells += 1
    return h.hexdigest(), n_cells


# ----------------------------------------------------------------------------- extraction
def formula_cache_state(path, wb_valued):
    """
    สถานะค่า cache ของสูตร — สำคัญมากต่อการประเมิน validator
    ไฟล์ export ดิบมักไม่มีค่า cache (openpyxl data_only อ่านได้ None)
    ส่วนไฟล์ที่เคยเปิดด้วย Excel แล้ว save จะมีค่า cache ครบ
    ⇒ กฎเดียวกันจะให้ผลต่างกันมากระหว่างสองสถานะนี้ ห้ามนำมาเทียบคะแนนกันตรง ๆ
    """
    wbf = load_workbook(path, read_only=True, data_only=False)
    coords = []
    for name in wbf.sheetnames:
        ws = wbf[name]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    coords.append((name, c.coordinate))
    wbf.close()
    n_formula = len(coords)
    if n_formula == 0:
        return 0, 0, "no_formulas"
    cached = 0
    by_sheet = {}
    for n, co in coords:
        by_sheet.setdefault(n, set()).add(co)
    for name, want in by_sheet.items():
        ws = wb_valued[name]
        for row in ws.iter_rows():
            for c in row:
                # read_only mode คืน EmptyCell ที่ไม่มี .coordinate — ต้องกรองด้วย value ก่อน
                if c.value is None:
                    continue
                if getattr(c, "coordinate", None) in want:
                    cached += 1
    if cached == 0:
        state = "no_cached_values"
    elif cached >= n_formula:
        state = "fully_cached"
    else:
        state = "partially_cached"
    return n_formula, cached, state


def head_scan(wb, max_row=8, max_col=30):
    """คืน [(sheet, coord, text)] ของเซลล์ข้อความในหัวไฟล์ — ใช้เป็นแหล่ง evidence"""
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
            for c in row:
                if isinstance(c.value, str) and c.value.strip():
                    out.append((name, c.coordinate, c.value.strip().replace("\n", " ")))
    return out


RE_STAGE = re.compile(r"ขั้น\s*(\d\.\d)\s*([^|]*)")
RE_YEAR_SYS = re.compile(r"ปีงบประมาณ\s*[:：]\s*(25\d\d)")
RE_YEAR_TITLE = re.compile(r"ประจำปีงบประมาณ\s*พ\.ศ\.\s*(25\d\d)")
RE_MINISTRY = re.compile(r"กระทรวง\s*[:：]\s*(\d{5})\s*([^|]+)")
RE_DEPT_CODE = re.compile(r"กรม\s*[:：]\s*(\d{5})\s*([^|]+)")
RE_DEPT_NAME = re.compile(r"กรม\s*[:：]\s*([^|\d][^|]*)")
RE_UNIT = re.compile(r"หน่วยปฏิบัติ\s*[:：]\s*(\d{6,})")


def extract_identity(cells):
    """ดึง identity จากเนื้อไฟล์ พร้อม evidence pointer ทุกฟิลด์"""
    f = {
        "fiscal_year_from_content": None,
        "ministry_code_from_content": None,
        "ministry_name_from_content": None,
        "agency_code_from_content": None,
        "agency_name_from_content": None,
        "agency_code_resolution": "unresolved",
        "operating_unit_from_content": None,
        "workflow_stage_from_content": None,
        "workflow_stage_label_raw": None,
    }
    ev = {}

    def put(key, val, sheet, coord, raw):
        if f.get(key) is None and val is not None:
            f[key] = val
            ev[key] = {"sheet": sheet, "cell": coord, "raw": raw[:300]}

    # รอบที่ 1 — บรรทัดระบบ (Bis1010) ให้ความสำคัญสูงสุด
    for sheet, coord, txt in cells:
        if "ปีงบประมาณ" in txt and "ขั้น" in txt:
            m = RE_YEAR_SYS.search(txt)
            if m:
                put("fiscal_year_from_content", int(m.group(1)), sheet, coord, txt)
            m = RE_STAGE.search(txt)
            if m:
                put("workflow_stage_from_content",
                    STAGE_ENUM.get(m.group(1), "stage_" + m.group(1)), sheet, coord, txt)
                put("workflow_stage_label_raw", ("ขั้น " + m.group(1) + " " + m.group(2)).strip(),
                    sheet, coord, txt)
            m = RE_MINISTRY.search(txt)
            if m:
                put("ministry_code_from_content", m.group(1), sheet, coord, txt)
                put("ministry_name_from_content", m.group(2).strip(), sheet, coord, txt)
            m = RE_DEPT_CODE.search(txt)
            if m:
                put("agency_code_from_content", m.group(1), sheet, coord, txt)
                put("agency_name_from_content", m.group(2).strip(), sheet, coord, txt)
                f["agency_code_resolution"] = "from_content_literal"
            else:
                m = RE_DEPT_NAME.search(txt)
                if m:
                    put("agency_name_from_content", m.group(1).strip(), sheet, coord, txt)
            m = RE_UNIT.search(txt)
            if m:
                put("operating_unit_from_content", m.group(1), sheet, coord, txt)

    # รอบที่ 2 — README / ชีต 'อ่านก่อน & ตรวจสอบ' ของไฟล์ผลลัพธ์ P1
    # บางไฟล์รวมหลาย key ไว้ในเซลล์เดียว คั่นด้วยช่องว่างซ้อน เช่น
    #   "หน่วยงาน: <ชื่อหน่วยงาน>   ปีงบประมาณ: <ปี>   หน่วย: ล้านบาท"
    # จึงต้องตัดที่ช่องว่าง 2 ตัวขึ้นไป ไม่งั้นชื่อหน่วยงานจะพ่วงข้อความอื่นมาด้วย
    def first_field(s):
        return re.split(r"\s{2,}", s.strip())[0].strip()

    for sheet, coord, txt in cells:
        m = re.match(r"^หน่วยงาน\s*[:：]\s*(.+)$", txt)
        if m:
            put("agency_name_from_content", first_field(m.group(1)), sheet, coord, txt)
        m = re.match(r"^กระทรวง\s*[:：]\s*(.+)$", txt)
        if m:
            put("ministry_name_from_content", first_field(m.group(1)), sheet, coord, txt)
        m = re.match(r"^ปีงบประมาณ\s*[:：]\s*(25\d\d)", txt)
        if m:
            put("fiscal_year_from_content", int(m.group(1)), sheet, coord, txt)

    # รอบที่ 3 — ตาราง key/value แนวนอน (ชีต 'อ่านก่อน & ตรวจสอบ': A='หน่วยงาน', B=ค่า)
    by_sheet = {}
    for sheet, coord, txt in cells:
        by_sheet.setdefault(sheet, {})[coord] = txt
    for sheet, m in by_sheet.items():
        for coord, txt in m.items():
            if txt in ("หน่วยงาน", "กระทรวง", "ปีงบประมาณ") and coord.startswith("A"):
                nb = "B" + coord[1:]
                val = m.get(nb)
                if not val:
                    continue
                if txt == "หน่วยงาน":
                    put("agency_name_from_content", val, sheet, nb, f"{txt} | {val}")
                elif txt == "กระทรวง":
                    put("ministry_name_from_content", val, sheet, nb, f"{txt} | {val}")
                elif txt == "ปีงบประมาณ" and re.match(r"^25\d\d$", val):
                    put("fiscal_year_from_content", int(val), sheet, nb, f"{txt} | {val}")

    # รอบที่ 4 — หัวรายงานแบบเก่า (ไม่มีบรรทัดขั้น): 'ประจำปีงบประมาณ พ.ศ. 25xx'
    for sheet, coord, txt in cells:
        m = RE_YEAR_TITLE.search(txt)
        if m:
            put("fiscal_year_from_content", int(m.group(1)), sheet, coord, txt)

    # resolve รหัสหน่วยงานจากชื่อ เมื่อเนื้อไฟล์ไม่ได้ให้รหัสมาตรง ๆ
    if f["agency_code_from_content"] is None and f["agency_name_from_content"]:
        code = AGENCY_NAME_TO_CODE.get(f["agency_name_from_content"].strip())
        if code:
            f["agency_code_from_content"] = code
            f["agency_code_resolution"] = "from_name_lookup"
            ev["agency_code_from_content"] = dict(
                ev.get("agency_name_from_content", {}),
                note="resolved via controlled vocabulary from agency_name_from_content",
            )
    if f["agency_code_from_content"] is None and f["operating_unit_from_content"]:
        # หน่วยปฏิบัติ 12 หลัก ขึ้นต้นด้วยรหัสกรม 5 หลัก
        cand = f["operating_unit_from_content"][:5]
        f["agency_code_from_content"] = cand
        f["agency_code_resolution"] = "from_operating_unit_prefix"
        ev["agency_code_from_content"] = dict(ev.get("operating_unit_from_content", {}),
                                              note="prefix of หน่วยปฏิบัติ")
    return f, ev


def detect_schema_version(sheetnames):
    s = set(sheetnames)
    n = len(sheetnames)
    if {"ภาพรวมประจำ-ลงทุน", "แผนยุทธฯ", "mask_cap"} <= s:
        return "BIS1010/13SHEET"
    if {"README", "Sheet4_สรุป", "MAD_Bridge", "Upload_Ready", "Cross_Check"} <= s:
        return ("P1OUT/6SHEET/LEGACY-S4-LAST"
                if sheetnames[-1] == "Sheet4_สรุป" else "P1OUT/6SHEET/README-FIRST")
    if {"อ่านก่อน & ตรวจสอบ", "Sheet4_สรุป", "Sheet5_รายละเอียด", "บัญชีและคำชี้แจง"} <= s:
        return "P1OUT/4SHEET/PRECHECK"
    if {"README", "อ่านก่อน & ตรวจสอบ", "Sheet4_สรุป"} <= s:
        return "SYN/P1OUT/5SHEET/CASE"
    if {"README", "MAD_Bridge", "Upload_Ready", "Cross_Check"} <= s:
        return "SYN/P1OUT/5SHEET/TEST"
    if s == {"แผนยุทธฯ", "แผนพื้นฐานฯ", "แผนบุคลากรฯ"}:
        return "SYN/BIS1010-MINI/3SHEET"
    if {"Summary", "Findings"} <= s:
        return "VALOUT/FINDINGS/2SHEET"
    if "00_README" in s:
        return "RESEARCH-INSTRUMENT/MULTISHEET"
    if s == {"P1"}:
        return "MOCK/P1-FLAT/1SHEET"
    if {"README", "P1"} <= s:
        return "MOCK/P1-FLAT/2SHEET"
    if {"รายละเอียด", "สรุป"} <= s:
        return "MOCK/P1-VALIDATOR-FIXTURE/3SHEET"
    if {"README", "บัญชีและคำชี้แจง"} <= s:
        return "SYN/P1-ACCOUNTS/2SHEET"
    return f"UNCLASSIFIED/{n}SHEET"


RE_FN_YEAR = re.compile(r"FY(\d\d)(?!\d)")
RE_FN_YEAR_FULL = re.compile(r"(25\d\d)")
RE_FN_AGENCY = re.compile(r"_(\d{5})_")


def parse_filename(basename: str):
    d = {"fiscal_year_from_filename": None, "agency_code_from_filename": None}
    m = RE_FN_YEAR.search(basename)
    if m:
        d["fiscal_year_from_filename"] = 2500 + int(m.group(1))
    else:
        m = RE_FN_YEAR_FULL.search(basename)
        if m:
            d["fiscal_year_from_filename"] = int(m.group(1))
    m = RE_FN_AGENCY.search(basename)
    if m:
        d["agency_code_from_filename"] = m.group(1)
    return d


def identity_agreement(content, fname):
    checks = []
    for a, b, label in (
        ("agency_code_from_content", "agency_code_from_filename", "agency"),
        ("fiscal_year_from_content", "fiscal_year_from_filename", "fiscal_year"),
    ):
        cv, fv = content.get(a), fname.get(b)
        if fv is None:
            checks.append((label, "filename_silent"))
        elif cv is None:
            checks.append((label, "content_silent"))
        elif str(cv) == str(fv):
            checks.append((label, "match"))
        else:
            checks.append((label, f"MISMATCH content={cv} filename={fv}"))
    states = [c[1] for c in checks]
    if any(s.startswith("MISMATCH") for s in states):
        verdict = "mismatch"
    elif all(s == "match" for s in states):
        verdict = "match"
    elif all(s in ("filename_silent", "content_silent") for s in states):
        verdict = "undetermined"
    else:
        verdict = "partial_match"
    return verdict, "; ".join(f"{k}={v}" for k, v in checks)


# ----------------------------------------------------------------------------- main
def main():
    targets = []
    for dirpath, dirnames, filenames in os.walk(ROOT):
        dirnames[:] = [d for d in dirnames if d not in (".git", os.path.basename(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))]
        for fn in filenames:
            if fn.lower().endswith((".xlsx", ".xlsm", ".xls")):
                targets.append(os.path.join(dirpath, fn))
    targets.sort()

    records = []
    for i, path in enumerate(targets, 1):
        rel = os.path.relpath(path, ROOT).replace("\\", "/")
        base = os.path.basename(path)
        st = os.stat(path)
        rec = {
            "record_id": f"P1REG-{i:04d}",
            "registry_schema_version": SCHEMA_VERSION,
            "observed_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "source_path": rel,
            "file_name": base,
            "file_size_bytes": st.st_size,
            "file_mtime_utc": datetime.fromtimestamp(st.st_mtime, timezone.utc)
                                      .isoformat(timespec="seconds"),
            "content_hash_algo": "sha256:file-bytes",
            "content_hash": file_bytes_sha256(path),
            "semantic_hash_algo": "sha256:sheetname+coord+normalized-value(data_only)",
            "semantic_content_hash": None,
            "semantic_hash_cell_count": None,
            "read_only_verified": True,
        }
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            rec["sheet_names"] = list(wb.sheetnames)
            rec["sheet_count"] = len(wb.sheetnames)
            rec["schema_version"] = detect_schema_version(wb.sheetnames)
            sh, nc = semantic_sha256(wb)
            rec["semantic_content_hash"] = sh
            rec["semantic_hash_cell_count"] = nc
            nf, nc_cached, fstate = formula_cache_state(path, wb)
            rec["formula_cell_count"] = nf
            rec["formula_cached_value_count"] = nc_cached
            rec["formula_cache_state"] = fstate
            cells = head_scan(wb)
            ident, ev = extract_identity(cells)
            rec.update(ident)
            rec["content_field_evidence"] = ev
            wb.close()
        except Exception as e:  # noqa: BLE001
            rec["extract_error"] = f"{type(e).__name__}: {e}"
            rec["schema_version"] = "UNREADABLE"

        rec.update(parse_filename(base))
        verdict, detail = identity_agreement(rec, rec)
        rec["identity_agreement"] = verdict
        rec["identity_agreement_detail"] = detail
        records.append(rec)

    # ---- ความสัมพันธ์: duplicate (ไบต์ตรง) / variant (semantic ตรงแต่ไบต์ต่าง)
    by_bytes, by_sem = {}, {}
    for r in records:
        by_bytes.setdefault(r["content_hash"], []).append(r)
        if r.get("semantic_content_hash"):
            by_sem.setdefault(r["semantic_content_hash"], []).append(r)
    for group in by_bytes.values():
        canon = group[0]["record_id"]
        for r in group:
            r["duplicate_of"] = None if r["record_id"] == canon else canon
            r["duplicate_group_size"] = len(group)
    for group in by_sem.values():
        distinct_bytes = {r["content_hash"] for r in group}
        canon = group[0]["record_id"]
        for r in group:
            r["variant_of"] = (canon if (len(distinct_bytes) > 1 and r["record_id"] != canon)
                               else None)
            r["semantic_group_size"] = len(group)

    # ---- declared_identity_key + identity_collision
    #      ป้ายที่ไฟล์ "ประกาศ" ว่าเป็นใคร/ปีไหน/สกุลไหน — ถ้าป้ายเดียวกันแต่ semantic hash ต่าง
    #      แปลว่ามีไฟล์มากกว่าหนึ่งชุดอ้างตัวเป็นสิ่งเดียวกัน = provenance ต้องตัดสินด้วยมนุษย์
    key_groups = {}
    for r in records:
        ac, fy = r.get("agency_code_from_content"), r.get("fiscal_year_from_content")
        if ac and fy:
            k = f"{ac}|{fy}|{r.get('schema_version')}"
        else:
            k = None          # ระบุตัวตนไม่ครบ → ไม่นำมาตรวจ collision
        r["declared_identity_key"] = k
        r["identity_collision"] = False
        r["identity_collision_peers"] = []
        if k:
            key_groups.setdefault(k, []).append(r)
    for k, g in key_groups.items():
        sems = {r.get("semantic_content_hash") for r in g}
        if len(sems) > 1:
            for r in g:
                r["identity_collision"] = True
                r["identity_collision_peers"] = sorted(
                    x["record_id"] for x in g if x["record_id"] != r["record_id"])

    # ---- scope / contamination (อิงหลักฐานเท่านั้น — ที่เหลือปล่อยเป็น unknown ให้คนตัดสิน)
    for r in records:
        code = r.get("agency_code_from_content")
        r["in_research_scope"] = bool(code and code in IN_SCOPE_AGENCIES)
        if r["file_name"] == MAPPING_DONOR_BASENAME:
            r["contamination"] = "mapping_donor"
            r["contamination_evidence"] = MAPPING_DONOR_EVIDENCE
        else:
            r["contamination"] = "unassigned"
            r["contamination_evidence"] = None
        # ฟิลด์ที่ต้องให้มนุษย์ตัดสิน — สคริปต์ไม่เดาให้
        r["data_role"] = "UNASSIGNED"
        r["role_rationale"] = None
        r["ground_truth_status"] = "none"
        r["ground_truth_source"] = None
        r["answer_key_ref"] = None
        r["seal_status"] = "open"
        r["sealed_at"] = None
        r["unseal_events"] = []
        r["derived_from"] = None
        r["supersedes"] = None
        r["superseded_by"] = None

    # ---- sensitivity / publishability
    for r in records:
        p = r["source_path"]
        syn = ("SYN/" in (r.get("schema_version") or "") or "MOCK/" in (r.get("schema_version") or "")
               or "เสมือน" in r["file_name"] or "SYNTHETIC" in r["file_name"].upper()
               or "synthetic" in p or "mock" in r["file_name"].lower())
        if (r.get("schema_version") or "").startswith("RESEARCH-INSTRUMENT"):
            r["sensitivity"] = "research_instrument"
        elif (r.get("schema_version") or "").startswith("VALOUT"):
            r["sensitivity"] = ("tool_output_from_synthetic" if "synthetic" in p or "ADVISOR" in p
                                else "tool_output_from_real")
        elif syn:
            r["sensitivity"] = "synthetic"
        else:
            r["sensitivity"] = "real_agency_data"
        r["publishable_to_github"] = r["sensitivity"] in ("synthetic", "research_instrument")

    # ---- pseudonym สำหรับมุมมองสาธารณะ
    codes = sorted({r["agency_code_from_content"] for r in records if r.get("agency_code_from_content")})
    pseudo = {c: f"A{i}" for i, c in enumerate(codes, 1)}
    for r in records:
        c = r.get("agency_code_from_content")
        r["agency_pseudonym"] = pseudo.get(c) if c else None

    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "registry.internal.json"), "w", encoding="utf-8") as f:
        json.dump({"registry_schema_version": SCHEMA_VERSION,
                   "generated_at_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                   "project_root": ROOT,
                   "status": "DRAFT — schema ยังไม่ได้รับอนุมัติ · data_role ทุกรายการยัง UNASSIGNED",
                   "record_count": len(records),
                   "agency_pseudonym_map": pseudo,
                   "records": records}, f, ensure_ascii=False, indent=2)

    internal_cols = ["record_id", "content_hash", "semantic_content_hash", "fiscal_year_from_content",
                     "agency_code_from_content", "agency_name_from_content", "agency_code_resolution",
                     "workflow_stage_from_content", "schema_version", "source_path",
                     "fiscal_year_from_filename", "agency_code_from_filename", "identity_agreement",
                     "identity_agreement_detail", "duplicate_of", "variant_of", "data_role",
                     "ground_truth_status", "in_research_scope", "contamination", "sensitivity",
                     "publishable_to_github", "seal_status", "file_size_bytes",
                     "formula_cache_state", "formula_cell_count",
                     "declared_identity_key", "identity_collision"]
    with open(os.path.join(OUT_DIR, "registry.internal.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=internal_cols, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in internal_cols})

    public_cols = ["record_id", "content_hash", "semantic_content_hash", "fiscal_year_from_content",
                   "agency_pseudonym", "workflow_stage_from_content", "schema_version",
                   "identity_agreement", "duplicate_of", "variant_of", "data_role",
                   "ground_truth_status", "in_research_scope", "contamination", "sensitivity",
                   "file_size_bytes", "formula_cache_state", "identity_collision"]
    with open(os.path.join(OUT_DIR, "registry.public.csv"), "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=public_cols, extrasaction="ignore")
        w.writeheader()
        for r in records:
            w.writerow({k: ("" if r.get(k) is None else r.get(k)) for k in public_cols})

    print(f"records                     : {len(records)}")
    print(f"distinct content_hash       : {len(by_bytes)}")
    print(f"distinct semantic hash      : {len(by_sem)}")
    print(f"identity mismatch           : {sum(1 for r in records if r['identity_agreement']=='mismatch')}")
    print(f"semantic-equal/byte-different: "
          f"{sum(1 for r in records if r.get('variant_of'))}")
    print(f"real_agency_data            : {sum(1 for r in records if r['sensitivity']=='real_agency_data')}")
    print(f"identity_collision groups   : "
          f"{len({r[chr(39)+chr(39)] if False else r['declared_identity_key'] for r in records if r['identity_collision']})}")
    from collections import Counter
    print("formula_cache_state         :", dict(Counter(r.get('formula_cache_state','?') for r in records)))
    print(f"written to                  : {OUT_DIR}")


if __name__ == "__main__":
    main()
